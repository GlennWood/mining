from __future__ import print_function
import re
import os
import sys
import jprops
import subprocess
import socket
import six
import yaml
import warnings
from openpyxl import load_workbook

try:
    from collections import ChainMap
except:
    # Somebody goofed when they deployed ChainMap for Python2.
    import importlib
    importlib.import_module('chainmap')

class Config(object):

    MINERS_XLSX = '/opt/mining/miners.xlsx'
    ALL_MEANS_ONCE = -11111
    I_AM_FORK = False
    RC_MAIN = 0

    # Global state from command-line options
    DRYRUN   = False
    VERBOSE  = False
    QUICK    = False
    QUERY    = False
    SCOPE    = False
    WIDE_OUT = False
    FORCE    = False
    PIP      = 'pip2'
    HOSTNAME = None#os.getenv('HOSTNAME','').upper()
    arguments = {}

    SHEETS = {
        'Globals': None,
        'CoinMiners': None,
        'Clients': None,
        'WhatToMine': None,
        'Overclock': None
        }
    
    GLOBALS = {}

    def __init__(self, argumentsIn):
        self.arguments = argumentsIn

        # Global Variables
        self.PLAT_COINS = {'AMD': {}, 'NVI': {}, 'BTH': {}}
        self.RIG = '' # just the first letter, capitalized, of the rig-Name
        self.WORKER_NAME = ''
        self.StatsUrls = {}
        self.ConvertUrls = {}
        
        if six.PY3: self.PIP = 'pip3'
        if not self.HOSTNAME: # So, HOSTNAME was not exported
            with open('/etc/hostname', 'r') as f: self.HOSTNAME = f.read().strip().upper()

        if self.arguments:
            # Command line options, values
            self.OPS = self.arguments['OPERATION'].split(',')
            if '--platform' in self.arguments and self.arguments['--platform'] is not None:
                self.PLATFORM = self.arguments['--platform']
            else:
                self.PLATFORM = os.getenv('PLATFORM','BTH')
            if not self.PLATFORM:
                self.PLATFORM = 'BTH'
            if '--url-port' in self.arguments: self.URL_PORT = self.arguments['--url-port']
            self.ALL_COINS = 'COIN' in self.arguments and ( self.arguments['COIN'] is None or len(self.arguments['COIN']) == 0 )
    
            # Command line options, booleans
            self.URL_PORT = self.SCOPE = None
            self.VERBOSE  = '-v'       in self.arguments and self.arguments['-v']
            self.PRINT    = '--print'  in self.arguments and self.arguments['--print']
            self.DRYRUN   = '--dryrun' in self.arguments and self.arguments['--dryrun']
            self.QUICK    = '--quick'  in self.arguments and self.arguments['--quick']
            self.QUERY    = '--query'  in self.arguments and self.arguments['--query']
            self.SCOPE    = '--scope'  in self.arguments and self.arguments['--scope']
            self.WIDE_OUT = '-l'       in self.arguments and self.arguments['-l']
            self.FORCE    = '--force'  in self.arguments and self.arguments['--force']
        else:
            self.arguments = {}
            self.URL_PORT = self.SCOPE = None
            self.VERBOSE  = False
            self.PRINT    = False
            self.DRYRUN   = False
            self.QUICK    = False
            self.QUERY    = False
            self.SCOPE    = None
            self.WIDE_OUT = False
            self.FORCE    = False
            self.ALL_COINS = True

        self.setup_config_dicts()
        self.setup_ansible_config()
        return
  

    # Parse sheets from miners.xslx
    def setup_config_dicts(self):

        # Put Stats sheet into stats_dict
        warnings.simplefilter("ignore")
        workbook = load_workbook(filename = self.MINERS_XLSX)#, read_only=True)
        warnings.resetwarnings()

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            self.SHEETS[sheet_name] = {}
            sheetNcols = sheet.max_column
            try:
                keys = [sheet.cell(1, col_index).value for col_index in range(1,sheetNcols)]
            except IndexError:
                sheetNcols = len(keys)
                if self.VERBOSE:
                    print('IndexError(xlrd): '+sheet_name+'[0]'+'.ncols='+str(sheet.ncols)+' >= '+str(sheetNcols))
            prev_key = None
            for row_index in range(2, sheet.max_row):
                try:
                    row = {}
                    for col_index in range(1,sheetNcols):
                        row[keys[col_index-1]]=sheet.cell(row_index,col_index)
                except IndexError:
                    if self.VERBOSE:
                        print('IndexError(xlrd): '+sheet_name+'['+str(row_index)+']'+'.ncols='+str(sheet.ncols)+' >= '+str(len(row)+1))
                if sheet_name == 'Globals': # Put Globals into substitution context
                    key = row['KEY']
                    if key and key.value and key.value.find('-') < 0:
                        self.GLOBALS[key] = str(row['VALUE'])
                elif sheet_name == 'CoinMiners': # CoinMiners' sheet is handled differently
                    row, prev_key = self.setup_CoinMiners_dict(row, prev_key)
                elif sheet_name == 'WhatToMine': # WhatToMine is keyed by column, not row
                    self.SHEETS[sheet_name] = self.pivot_sheet(self.SHEETS[sheet_name], row, keys)
                else:
                    key = row[keys[0]]
                    if key and key.value:
                        self.SHEETS[sheet_name][key.value.upper()] = row

        if self.ALL_COINS: 
            self.arguments['COIN'] = [x.upper() for x in sorted(list(self.SHEETS['CoinMiners'].keys()))]   

        return [self.SHEETS['Stats'],self.StatsUrls,self.ConvertUrls,self.SHEETS['CoinMiners'],self.SHEETS['Clients']]

    ''' The miners.xslx/WhatToMine spreadsheet is mapped from the keys on the first
    row into the columns under each key, so we do that way here rather than the 
    default of first-column => columns (as above).
    '''
    def pivot_sheet(self, sheet, row, keys):
        for key in keys:
            if key == 'Param':
                param = row[key].value
            else:
                row_val = row[key].value
                if key not in sheet:
                    sheet[key] = { }
                sheet[key][param] = row_val
        return sheet


    # The CoinMiners sheet enables a number of special configurable configurating.
    #   1. The OPTIONS field is configured on a line following the main config (to give more room to spell out the OPTIONS)
    #   2. The MINER can be specified as a key to the Clients spreadsheet to spell out the CoinMiner's options
    #   2a. (this MNEMONICS capability is actually facilitated in the start option-module)
    #
    def setup_CoinMiners_dict(self, row, prev_key):
        if prev_key is None and row['COIN'] and row['COIN'].value.strip():
            prev_key = row['COIN'].value.upper()
        # The CoinMiners sheet has a special way of enabling a wider OPTIONS cell
        row['OPTIONS']='' # There is no OPTIONS column native to the CoinMiners sheet
        if row['COIN'].value and not row['COIN'].value.strip():
            self.SHEETS['CoinMiners'][prev_key]['OPTIONS'] = row['MINER']
            return row, prev_key
        # Apply the '--url-port' command-line option, if exists.
        if self.URL_PORT:
            row['URL_PORT'] = self.URL_PORT
        # Provision $URL and $PORT as alternatives to $URL_PORT
        regex = re.compile(r'(.*)[:]([0-9]{4,5})')
        match = regex.match(row['URL_PORT'].value) if row['URL_PORT'].value else None
        if match != None:
            row['URL'] = match.group(1)
            row['PORT'] = match.group(2)
        # Replace $USER_PSW, and/or $USER and $PSW, with configured value(s)
        user_psw = row['USER_PSW'].value if  row['USER_PSW'].value else ""
        USER_PSW = user_psw.split(':')
        if len(USER_PSW) > 1:
            row['USER'] = USER_PSW[0]
            row['PSW']  = USER_PSW[1]
        else:
            row['USER'] = row['USER_PSW']
            row['PSW']  = ''

        prev_key = row['COIN'].value.upper() if row['COIN'].value else ""

        #if prev_key: 
        # Index each coinMiner under the applicable platform, AMD, NVI or BTH
        plat = row['PLAT'].value
        if plat is None or plat == '': plat = 'BTH'
        if plat == 'BTH':
            for key in self.PLAT_COINS:
                self.PLAT_COINS[key][prev_key] = row
        else:
            self.PLAT_COINS[plat][prev_key] = row
        # TODO That one is deprecated since we want to use the PLATFORM specific version every time ...
        #   ... but this will have to wait for corrective coding later on where yhat dict is used.
        self.SHEETS['CoinMiners'][prev_key] = row

        return row, prev_key

    # Configurations from ansible
    def setup_ansible_config(self):
        ansibleFilename = '/etc/ansible/hosts'
        self.ANSIBLE_HOSTS = { }
   
        try:
            # jprops does a good job of handling the minutiae of parsing a properties file, but
            #   it does not handle sections. So we use a hacky wrapper on jprops to do that.
            lines = open(ansibleFilename).read().splitlines()
            fo = sectionFilename = None
            sectionFiles = []
            for line in lines:
                regex = re.compile(r'[[]([^]]*)[]]')
                match = regex.match(line)
                if match:
                    section = match.group(1)
                    if fo: fo.close()
                    sectionFilename = '/var/local/ramdisk/ansible.'+section+'.ini'
                    fo = open(sectionFilename, 'w')
                    sectionFiles.append(sectionFilename)
                else:
                    if fo: fo.write(line+"\n")
            if fo: fo.close()

            ansibleMinersFilename = '/var/local/ramdisk/ansible.miners.ini'
            with open(ansibleMinersFilename) as fh:
                for ip, value in jprops.iter_properties(fh):
                    (hostname, platform) = value.strip().split(' ')
                    hostname = hostname.split('=')[1]
                    platform = platform.split('=')[1]
                    self.ANSIBLE_HOSTS[hostname] = {'hostname': hostname, 'platform': platform, 'ip': ip}

            # We don't need these section files anymore.
            for fn in sectionFiles: os.remove(fn)

        except IOError as ex:
            print(str(ex))
        except AttributeError as ex:
            print(ansibleFilename+' format is invalid (for miners).' + str(ex))
        except:
            print ( sys.exc_info()[0] )

    ### ###########################################################
    @staticmethod
    def load_sources_yml():
        with open("/opt/mining/conf/sources.yml", 'r') as stream:
            try:
                SOURCES_X = yaml.load(stream, Loader=yaml.FullLoader)
            except yaml.YAMLError as exc:
                print(exc)
                sys.exit(1)
        return SOURCES_X
    
    # Find the given ticker in the CoinMiners table for this PLATFORM
    #   Return None if not found
    #   If verbose=True: print error message if not found
    def findTickerInPlatformCoinMiners(self, ticker, verbose=False):
        if ticker and ticker in self.PLAT_COINS[self.PLATFORM]:
            return self.PLAT_COINS[self.PLATFORM][ticker]
        if self.PLATFORM == 'BTH':
            return self.findTickerInCoinMiners(ticker, verbose)
        if verbose: print("Coin '" + ticker + "' is not configured for this platform='"+self.PLATFORM+"'.", file=sys.stderr)
        return None

    # Find the given ticker in the CoinMiners table, regardless of PLATFORM
    #   Return None if not found
    #   If verbose=True: print error message if not found
    def findTickerInCoinMiners(self, ticker, verbose=False):
        if ticker and ticker in self.SHEETS['CoinMiners']:
            return self.SHEETS['CoinMiners'][ticker]
        if verbose: print("Coin '" + ticker + "' is not configured in miners.xslx/CoinMiners.", file=sys.stderr)
        return None

    def substitute(self, row_id, arg):
        rslt = arg
        context = ChainMap({'COIN': self.SHEETS['CoinMiners'][row_id]['COIN'].split('-',1)[0]},
                            self.SHEETS['CoinMiners'][row_id], os.environ, self.GLOBALS)
        # Substitute in reverse sequence of name's length
        # (I don't know who is in charge of conceptual integrity of the Python language. Anyone?)
        keys = sorted(context.keys(), key=len)
        keys.reverse() # list(keys).sort(key=lambda item: (-len(item), item))
        for trans in range(0,2): # loop thrice to enable transitive substitutions
            if trans>9:
                print(rslt)
            for key in keys:
                val = context[key]
                if val == '': continue
                # Substitute both '$NAME" and '<NAME>'
                rslt = rslt.replace('$'+key,val).replace('<'+key+'>',val)
        return rslt

    def get_sttyDims(self):
        try:
            sttyRow, sttyCol = subprocess.check_output(['stty', 'size'], stderr=subprocess.PIPE).split()
        except:
            return int(99), int(999)
        return int(sttyRow), int(sttyCol)
    
    def get_sttyColumnsMaxRigNameLen(self):
        maxRigNameLen = max(self.ANSIBLE_HOSTS.keys(), key = len)
        return self.get_sttyDims()[1], len(maxRigNameLen)

    # WORKER_NAME varies by the coin we are processing, so this
    #   method is a convenient central point for set/getting it.
    def workerName(self,ticker):
        if not self.RIG:
            self.RIG = socket.gethostname().replace('rig-','')[0].upper()
            os.environ['RIG'] = self.RIG
        if ticker:
            self.WORKER_NAME = self.RIG + '-' + ticker + '-miner'
        return self.WORKER_NAME
